#!/usr/bin/env python
# -*- coding: utf-8 -*-
import csv

import os
from pprint import pprint
import openpyxl
from openpyxl import Workbook



path_79 = "d:\счета new\питон\ексель маржа"
os.chdir(path_79)
m = os.listdir(path= path_79 )
# file = path_79 + '\' + m[0]
m=m[0]

list_full=[]
with open(m, newline='') as csvfile:
    spamreader = csv.reader(csvfile,  quotechar=',')
    for line in spamreader :

        temp = (",".join(line).split(';'))
        temp[0]=temp[0].split(" ")
        list_full.append(temp)

# print(list_full[1808][0][0], 2+2)
# print(list_full[1808][6], 2+2)
data_list = []

for line in list_full :

    if line[0][0] not in data_list  and   line[0][0] != 'CDR':
        data_list.append(line[0][0])

data_list.sort()
data_list_dict = data_list[:]


# создаем словарь из списка
for number in range(len(data_list_dict)):
    data_list_dict[number]={data_list_dict[number]:[]}

s = data_list_dict[0].items()
for i in s:
    key= i[0]


s= (data_list_dict[0].get(key))
# s.append(2)
mon=0
# добавляем значение опеаратора для ключей по датам
for line in range(len(data_list_dict)):
    s = data_list_dict[line].items()
    for i in s:
        key = i[0]

    mon = 0
    list_of_dates=(data_list_dict[line].get(key))  # список оперататоров в дате
    for full in list_full :
        mon=mon+1


        if key == full[0][0]   :

            # print (list_of_dates, "ок, сэр")
            # print (full[6], 'key', full[0][0])
            dict_oper ={full[6]:[0,0,0,0]}
            if dict_oper not in list_of_dates:

                list_of_dates.append(dict_oper)


# pprint (data_list_dict[0])
for line in range(len(data_list_dict)):
    s = data_list_dict[line].items()
    for i in s:
        key = i[0]
    # print (key)
    list_of_dates = (data_list_dict[line].get(key))  # список оперататоров в дате
    for item in list_of_dates:
           for some in item :
                # print (some,item[some])
                for xuk in list_full :
                    # print(xuk[0][0],key,some,xuk[6],type(xuk[7]))
                    if xuk[0][0] == key and some == xuk[6]:
                        item[some][0]=item[some][0]+float(xuk[9])
                        item[some][1] = item[some][1] + float(xuk[12])
                        item[some][2] = item[some][0] - item[some][1]
                        item[some][3] = item[some][3] + float(xuk[1])/60

counter=1

wb = Workbook()
ws = wb.active

#собственно вставляем значения из словаря
for line in range(len(data_list_dict)):
    s = data_list_dict[line].items()
    for i in s:
        key = i[0]
    # print (key)
    list_of_dates = (data_list_dict[line].get(key))  # список оперататоров в дате
    for item in list_of_dates:
        for some in item:
            # print (item)
             counter  = counter+1
             # print(item[some], counter)
             # print( counter)
             ws.cell(row=counter, column=1).value = key
             ws.cell(row=counter, column=2).value = some
             ws.cell(row=counter, column=3).value = round(item[some][0],2)
             ws.cell(row=counter, column=4).value = round(item[some][1],2)
             ws.cell(row=counter, column=5).value = round(item[some][2],2)
             ws.cell(row=counter, column=6).value = round(item[some][3],2)

# Шапка в екселе

ws.cell(row=1, column=1).value="date"
ws.cell(row=1, column=2).value="Vendor"
ws.cell(row=1, column=3).value="IN"
ws.cell(row=1, column=4).value="OUT"
ws.cell(row=1, column=5).value="margin"
ws.cell(row=1, column=6).value="time"
#сохраняем книгу
#todo сделать дату в названии файла , причесать файл
#
wb.save("result.xlsx")